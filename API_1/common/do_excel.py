import openpyxl

class Case:
    """测试用例类"""
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.method=None
        self.headers=None
        self.data=None
        self.expected=None
        self.sql=None
        self.actual=None
        self.result=None
class DoExcel:
    """读取和写入excel"""
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.wb=openpyxl.load_workbook(self.file_name)
        self.sheet_name=sheet_name
        self.sheet=self.wb[self.sheet_name]

    def get_cass(self):
        max_row=self.sheet.max_row #获取最大列数
        # print(max_row)
        cases=[]
        for row in range(2,max_row+1):
            case = Case()
            case.case_id=self.sheet.cell(row,1).value
            case.title=self.sheet.cell(row,2).value
            case.url=self.sheet.cell(row,3).value
            case.method=self.sheet.cell(row,4).value
            case.headers=self.sheet.cell(row,5).value
            case.data=self.sheet.cell(row,6).value
            case.expected=self.sheet.cell(row,7).value
            case.sql=self.sheet.cell(row,9).value
            cases.append(case)
            # print(cases)
        self.wb.close()
        return cases

    def write_excel(self,row,actual,result):
        self.sheet.cell(row,8).value=actual
        self.sheet.cell(row,9).value=result
        self.wb.save(self.file_name)
        self.wb.close()



if __name__=='__main__':
    ex=DoExcel('E:\workspace\API\API_1\data\cases.xlsx','hotLabel')
    l=ex.get_cass()
    # ex.write_excel(2,'测试','你好')
    print(l[0].title)



